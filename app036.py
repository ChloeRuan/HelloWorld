# pypi and pip ( people already build for other people to use)
# cell is a package, __init__ py is a module
# instead of openpyxl.** we can write xl.


import openpyxl as xl

wb = xl.load_workbook('transactions.xlsx')  # load_workbook is a function to load excel
sheet = wb['Sheet1']  # case sensitive

cell = sheet['a1']
"""
cell = sheet.cell(1, 1)
"""
print(cell.value)
#  sheet.cell(1,1 )
# how many rows
print(sheet.max_row)

# how to extract values of a specific column
"""
for row in range(2, sheet.max_row + 1): #why? if we write (1, 4), it will only genderate 1, ,2, 3
    cell = sheet.cell(row, 3)   # 3 stands for column
    print(cell.value)
"""

# correct the prices of the spread sheet
for row in range(2, sheet.max_row + 1):  # why? if we write (1, 4), it will only genderate 1, ,2, 3
    cell = sheet.cell(row, 3)  # 3 stands for column
    correct_price = cell.value * 0.9
    corrected_price_cell = sheet.cell(row, 4)
    corrected_price_cell.value = correct_price

"""
wb.save('transaction2.xlsx')
"""
from openpyxl.chart import BarChart, Reference

# create a chart
values = Reference(sheet, min_row=2, max_row=sheet.max_row, min_col=4, max_col=4)

chart = BarChart()
chart.add_data(values)
sheet.add_chart(chart, 'e2')  # where do we want to add this chart
